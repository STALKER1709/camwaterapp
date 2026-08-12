"""Gestion du fichier Excel général `CAMWATER_Pointage_General.xlsx`.

Garanties apportées par ce module :

* **Atomicité** — le classeur est chargé, modifié en mémoire, écrit dans un
  fichier temporaire du même dossier, puis basculé par `os.replace()`. Une
  coupure en cours d'écriture laisse donc toujours l'ancien fichier intact :
  jamais de ligne à moitié écrite.
* **Exclusion mutuelle** — un verrou fichier sérialise les écritures
  concurrentes venant de plusieurs postes ou de plusieurs requêtes HTTP.
* **Idempotence** — la feuille `Journal` conserve l'empreinte SHA-256 de chaque
  fichier traité, ce qui permet de refuser un doublon.
* **Traçabilité** — les feuilles `Résumé` et `Anomalies` sont reconstruites à
  chaque enregistrement à partir des données réelles du classeur.
"""

from __future__ import annotations

import logging
import os
import re
import shutil
import tempfile
import threading
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .config import (
    CLE_FACTURE,
    COLONNES,
    COLONNES_NUMERIQUES,
    COMPLETER_MOIS_MANQUANTS,
    EXCEL_LOCK_PATH,
    EXCEL_PATH,
    FEUILLE_ANOMALIES,
    FEUILLE_ECARTEES,
    FEUILLE_ECHECS,
    FEUILLE_JOURNAL,
    FEUILLE_RESUME,
    FEUILLE_UNIQUE,
    MARQUEUR_A_VERIFIER,
    MARQUEUR_ECARTEE,
    MARQUEUR_ILLISIBLE,
)
from .models import STATUT_OK, LigneFacture
from .periodes import mois_manquants, periode_vers_cle

logger = logging.getLogger(__name__)

__all__ = [
    "ETAT_INCOMPLET",
    "ETAT_MANQUANTE",
    "ETAT_RECUE",
    "ExcelError",
    "ExcelManager",
    "identites_facture",
    "nom_feuille",
]


class ExcelError(RuntimeError):
    """Le classeur n'a pas pu être lu ou écrit."""


# --------------------------------------------------------------------------- #
# Mise en forme
# --------------------------------------------------------------------------- #

_ENTETE_POLICE = Font(bold=True, color="FFFFFF", size=11)
_ENTETE_FOND = PatternFill("solid", fgColor="1F4E79")
_ENTETE_ALIGNEMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_FOND_A_VERIFIER = PatternFill("solid", fgColor="FFF2CC")
_FOND_ILLISIBLE = PatternFill("solid", fgColor="F8CBAD")
_FOND_ECARTEE = PatternFill("solid", fgColor="E7E6E6")
_FOND_MOIS_MANQUANT = PatternFill("solid", fgColor="FCE4EC")
#: Total sous-évalué : au moins un chiffre de l'administration n'a pas été lu.
_FOND_INCOMPLET = PatternFill("solid", fgColor="FFF2CC")
#: Facture déposée mais non lue : elle reste à retraiter.
_FOND_ECHEC = PatternFill("solid", fgColor="F8CBAD")
_FORMAT_NOMBRE = "#,##0"
_LARGEURS = {
    "DR": 14,
    "Agence": 10,
    "Période": 12,
    "Compte client": 18,
    "Nom du client": 34,
    "Ville": 16,
    "Code abonnement (PL)": 20,
    "Nom abonné": 34,
    "N° compteur": 16,
    "Ministère_2026": 18,
}

#: En-tête de la feuille Résumé. « État » distingue un mois effectivement
#: pointé d'un mois pour lequel aucune facture n'a été reçue.
_ENTETES_RESUME = (
    "Administration",
    "Période",
    "État",
    "Nb lignes",
    "Consommation (m³)",
    "Montant HT",
    "TVA",
    "Montant TTC",
    "Lignes à vérifier",
    "Lignes illisibles",
)

ETAT_RECUE = "Reçue"
ETAT_MANQUANTE = "MANQUANTE"

#: Une administration dont au moins un chiffre est illisible : les montants de
#: sa ligne de Résumé sont des **minorants**, jamais le total réel. Sans cette
#: mention, un total sous-évalué se lit comme un total définitif.
ETAT_INCOMPLET = "Reçue — total incomplet"

#: Colonnes de la feuille des factures en échec.
_ENTETES_ECHECS = (
    "Horodatage",
    "Fichier source",
    "Empreinte SHA-256",
    "Utilisateur",
    "Pages lues",
    "Motif de l'échec",
)

#: Colonnes techniques ajoutées après les 17 colonnes standard.
COLONNES_TECHNIQUES = ("Statut", "Fichier source", "Anomalies")

_CARACTERES_INTERDITS = re.compile(r"[\[\]:*?/\\]")
_verrou_processus = threading.Lock()


def nom_feuille(administration: str) -> str:
    """Transforme un nom d'administration en nom de feuille Excel valide."""
    nom = _CARACTERES_INTERDITS.sub("-", (administration or "").strip())
    nom = nom.replace("'", "").strip() or MARQUEUR_A_VERIFIER
    nom = _CARACTERES_INTERDITS.sub("-", nom)
    return nom[:31]


def identites_facture(lignes: Iterable[LigneFacture]) -> set[tuple[str, str]]:
    """Couples (compte client, période) portés par un lot de lignes.

    C'est l'identité **métier** d'une facture : un même compte facturé sur une
    même période est la même facture, quel que soit le fichier scanné qui la
    porte. Les lignes écartées (sans compte) n'ont pas d'identité.
    """
    identites: set[tuple[str, str]] = set()
    for ligne in lignes:
        if ligne.statut == MARQUEUR_ECARTEE:
            continue
        compte = str(ligne.valeurs.get(CLE_FACTURE[0]) or "").strip()
        periode = str(ligne.valeurs.get(CLE_FACTURE[1]) or "").strip()
        if compte and periode:
            identites.add((compte, periode))
    return identites


def _verrou_fichier():
    """Verrou inter-processus si `filelock` est disponible, sinon verrou local."""
    try:
        from filelock import FileLock

        EXCEL_LOCK_PATH.parent.mkdir(parents=True, exist_ok=True)
        return FileLock(str(EXCEL_LOCK_PATH), timeout=120)
    except ImportError:  # pragma: no cover - dépendance déclarée
        logger.warning(
            "Le paquet « filelock » est absent : la protection contre les écritures "
            "concurrentes est limitée au processus courant."
        )
        return _verrou_processus


class ExcelManager:
    """Point d'entrée unique pour lire et écrire le classeur de pointage."""

    def __init__(self, chemin: Optional[Path] = None) -> None:
        self.chemin = Path(chemin) if chemin else EXCEL_PATH
        self.chemin.parent.mkdir(parents=True, exist_ok=True)

    # -- Chargement / création ------------------------------------------- #

    def _nouveau_classeur(self) -> Workbook:
        classeur = Workbook()
        classeur.remove(classeur.active)
        self._creer_feuille_resume(classeur)
        self._creer_feuille_anomalies(classeur)
        self._creer_feuille_journal(classeur)
        logger.info("Nouveau classeur créé : %s", self.chemin)
        return classeur

    def _charger(self) -> Workbook:
        if not self.chemin.exists():
            return self._nouveau_classeur()
        try:
            return load_workbook(self.chemin)
        except Exception as exc:
            raise ExcelError(
                f"Impossible d'ouvrir « {self.chemin} » : {exc}. "
                "Fermez le fichier s'il est ouvert dans Excel, ou restaurez une sauvegarde."
            ) from exc

    def _enregistrer(self, classeur: Workbook) -> None:
        """Écriture transactionnelle : fichier temporaire puis bascule atomique."""
        dossier = self.chemin.parent
        descripteur, provisoire = tempfile.mkstemp(
            prefix=f".{self.chemin.stem}_", suffix=".tmp.xlsx", dir=str(dossier)
        )
        os.close(descripteur)
        provisoire_path = Path(provisoire)
        try:
            classeur.save(provisoire_path)
            if self.chemin.exists():
                # Sauvegarde de secours de la version précédente.
                shutil.copy2(self.chemin, self.chemin.with_suffix(".xlsx.bak"))
            os.replace(provisoire_path, self.chemin)
        except PermissionError as exc:
            # Cas le plus fréquent sous Windows : le classeur est ouvert dans
            # Excel, qui pose un verrou exclusif empêchant le remplacement.
            provisoire_path.unlink(missing_ok=True)
            raise ExcelError(
                f"« {self.chemin.name} » est verrouillé par une autre application "
                "(le plus souvent Excel, qui l'a ouvert). Fermez-le puis relancez le "
                "traitement : aucune ligne n'a été écrite, le fichier est intact."
            ) from exc
        except Exception as exc:
            provisoire_path.unlink(missing_ok=True)
            raise ExcelError(f"Échec de l'enregistrement de « {self.chemin} » : {exc}") from exc

    # -- Création des feuilles -------------------------------------------- #

    @staticmethod
    def _styliser_entete(feuille: Worksheet, entetes: Iterable[str]) -> None:
        entetes = list(entetes)
        feuille.append(entetes)
        for index, titre in enumerate(entetes, start=1):
            cellule = feuille.cell(row=1, column=index)
            cellule.font = _ENTETE_POLICE
            cellule.fill = _ENTETE_FOND
            cellule.alignment = _ENTETE_ALIGNEMENT
            feuille.column_dimensions[get_column_letter(index)].width = _LARGEURS.get(titre, 15)
        feuille.freeze_panes = "A2"
        feuille.auto_filter.ref = (
            f"A1:{get_column_letter(len(entetes))}1" if entetes else None
        )

    def _creer_feuille_donnees(self, classeur: Workbook, nom: str) -> Worksheet:
        feuille = classeur.create_sheet(nom)
        self._styliser_entete(feuille, list(COLONNES) + list(COLONNES_TECHNIQUES))
        return feuille

    def _creer_feuille_resume(self, classeur: Workbook) -> Worksheet:
        feuille = classeur.create_sheet(FEUILLE_RESUME, 0)
        self._styliser_entete(
            feuille,
            _ENTETES_RESUME,
        )
        return feuille

    def _creer_feuille_anomalies(self, classeur: Workbook) -> Worksheet:
        feuille = classeur.create_sheet(FEUILLE_ANOMALIES)
        self._styliser_entete(
            feuille,
            (
                "Fichier source",
                "Feuille",
                "Statut",
                "Compte client",
                "Code abonnement (PL)",
                "Nom abonné",
                "Anomalie",
            ),
        )
        return feuille

    def _creer_feuille_ecartees(self, classeur: Workbook) -> Worksheet:
        feuille = classeur.create_sheet(FEUILLE_ECARTEES)
        self._styliser_entete(feuille, list(COLONNES) + list(COLONNES_TECHNIQUES))
        return feuille

    def _creer_feuille_echecs(self, classeur: Workbook) -> Worksheet:
        feuille = classeur.create_sheet(FEUILLE_ECHECS)
        self._styliser_entete(feuille, _ENTETES_ECHECS)
        return feuille

    def _creer_feuille_journal(self, classeur: Workbook) -> Worksheet:
        feuille = classeur.create_sheet(FEUILLE_JOURNAL)
        self._styliser_entete(
            feuille,
            (
                "Horodatage",
                "Fichier source",
                "Empreinte SHA-256",
                "Utilisateur",
                "Compte client",
                "Période",
                "Pages",
                "Lignes écrites",
                "Lignes à vérifier",
                "Lignes illisibles",
                "Lignes écartées",
                "Confiance",
            ),
        )
        return feuille

    def _feuille(self, classeur: Workbook, nom: str) -> Worksheet:
        if nom in classeur.sheetnames:
            return classeur[nom]
        if nom == FEUILLE_RESUME:
            return self._creer_feuille_resume(classeur)
        if nom == FEUILLE_ANOMALIES:
            return self._creer_feuille_anomalies(classeur)
        if nom == FEUILLE_JOURNAL:
            return self._creer_feuille_journal(classeur)
        if nom == FEUILLE_ECARTEES:
            return self._creer_feuille_ecartees(classeur)
        if nom == FEUILLE_ECHECS:
            return self._creer_feuille_echecs(classeur)
        return self._creer_feuille_donnees(classeur, nom)

    @staticmethod
    def _feuilles_donnees(classeur: Workbook) -> list[Worksheet]:
        reservees = {
            FEUILLE_RESUME,
            FEUILLE_ANOMALIES,
            FEUILLE_JOURNAL,
            FEUILLE_ECARTEES,
            FEUILLE_ECHECS,
        }
        return [classeur[nom] for nom in classeur.sheetnames if nom not in reservees]

    # -- Écriture ---------------------------------------------------------- #

    @staticmethod
    def _ecrire_ligne(feuille: Worksheet, ligne: LigneFacture, fichier: str) -> None:
        valeurs: list[Any] = []
        for colonne in COLONNES:
            valeur = ligne.valeurs.get(colonne)
            if isinstance(valeur, Decimal):
                valeur = float(valeur)
            valeurs.append(valeur)
        valeurs.extend(
            [ligne.statut, fichier, " | ".join(ligne.anomalies) if ligne.anomalies else ""]
        )
        feuille.append(valeurs)

        numero_ligne = feuille.max_row
        for index, colonne in enumerate(COLONNES, start=1):
            if colonne in COLONNES_NUMERIQUES:
                cellule = feuille.cell(row=numero_ligne, column=index)
                if isinstance(cellule.value, (int, float)):
                    cellule.number_format = _FORMAT_NOMBRE

        fond = None
        if ligne.statut == MARQUEUR_ECARTEE:
            fond = _FOND_ECARTEE
        elif ligne.statut == MARQUEUR_ILLISIBLE:
            fond = _FOND_ILLISIBLE
        elif ligne.statut == MARQUEUR_A_VERIFIER:
            fond = _FOND_A_VERIFIER
        if fond is not None:
            for index in range(1, len(valeurs) + 1):
                feuille.cell(row=numero_ligne, column=index).fill = fond

    def _ajouter_anomalies(
        self, classeur: Workbook, lignes: list[LigneFacture], fichier: str
    ) -> None:
        feuille = self._feuille(classeur, FEUILLE_ANOMALIES)
        for ligne in lignes:
            if ligne.statut == STATUT_OK or not ligne.anomalies:
                continue
            for anomalie in ligne.anomalies:
                feuille.append(
                    [
                        fichier,
                        nom_feuille(FEUILLE_UNIQUE or ligne.administration),
                        ligne.statut,
                        ligne.valeurs.get("Compte client"),
                        ligne.valeurs.get("Code abonnement (PL)"),
                        ligne.valeurs.get("Nom abonné"),
                        anomalie,
                    ]
                )

    def _journaliser(
        self,
        classeur: Workbook,
        fichier: str,
        empreinte: str,
        utilisateur: str,
        pages: int,
        lignes: list[LigneFacture],
        confiance: float,
    ) -> None:
        feuille = self._feuille(classeur, FEUILLE_JOURNAL)
        identites = identites_facture(lignes)
        feuille.append(
            [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                fichier,
                empreinte,
                utilisateur or "-",
                " ; ".join(sorted({c for c, _ in identites})) or "-",
                " ; ".join(sorted({p for _, p in identites})) or "-",
                pages,
                len(lignes),
                sum(1 for l in lignes if l.statut == MARQUEUR_A_VERIFIER),
                sum(1 for l in lignes if l.statut == MARQUEUR_ILLISIBLE),
                sum(1 for l in lignes if l.statut == MARQUEUR_ECARTEE),
                round(confiance, 3),
            ]
        )

    def _reconstruire_resume(self, classeur: Workbook) -> None:
        """Recalcule intégralement la feuille Résumé depuis les feuilles de données."""
        if FEUILLE_RESUME in classeur.sheetnames:
            index = classeur.sheetnames.index(FEUILLE_RESUME)
            classeur.remove(classeur[FEUILLE_RESUME])
        else:
            index = 0
        feuille = classeur.create_sheet(FEUILLE_RESUME, index)
        self._styliser_entete(feuille, _ENTETES_RESUME)

        positions = {colonne: i for i, colonne in enumerate(COLONNES)}
        position_statut = len(COLONNES)
        agregats: dict[tuple[str, str], dict[str, float]] = {}

        for onglet in self._feuilles_donnees(classeur):
            for valeurs in onglet.iter_rows(min_row=2, values_only=True):
                if valeurs is None or all(v is None for v in valeurs):
                    continue
                # Une feuille plus étroite que les 17 colonnes standard (onglet
                # ajouté à la main, classeur d'un format antérieur) ne doit pas
                # interrompre le pointage : on la signale et on passe.
                if len(valeurs) <= positions["Montant TTC"]:
                    logger.warning(
                        "Feuille « %s » : ligne de %d colonne(s) au lieu de %d — ignorée.",
                        onglet.title,
                        len(valeurs),
                        len(COLONNES),
                    )
                    continue
                administration = str(valeurs[positions["Ministère_2026"]] or MARQUEUR_A_VERIFIER)
                periode = str(valeurs[positions["Période"]] or "-")
                statut = (
                    str(valeurs[position_statut])
                    if len(valeurs) > position_statut and valeurs[position_statut]
                    else STATUT_OK
                )
                agregat = agregats.setdefault(
                    (administration, periode),
                    {
                        "lignes": 0.0,
                        "consommation": 0.0,
                        "ht": 0.0,
                        "tva": 0.0,
                        "ttc": 0.0,
                        "a_verifier": 0.0,
                        "illisible": 0.0,
                        "non_chiffre": 0.0,
                    },
                )
                agregat["lignes"] += 1
                for cle, colonne in (
                    ("consommation", "Consommation"),
                    ("ht", "Montant HT"),
                    ("tva", "TVA"),
                    ("ttc", "Montant TTC"),
                ):
                    valeur = valeurs[positions[colonne]]
                    if isinstance(valeur, (int, float)):
                        agregat[cle] += float(valeur)
                    elif valeur is not None:
                        # Un chiffre non lu (« ILLISIBLE ») ne s'additionne pas :
                        # le total de cette administration devient un minorant.
                        agregat["non_chiffre"] += 1
                if statut == MARQUEUR_A_VERIFIER:
                    agregat["a_verifier"] += 1
                elif statut == MARQUEUR_ILLISIBLE:
                    agregat["illisible"] += 1

        # Compléter avec les mois sans facture : sur une étendue de plusieurs
        # mois, un mois absent du classeur est invisible sans cela. On borne
        # l'étendue par la plus ancienne et la plus récente période **du
        # classeur entier**, afin que chaque administration montre ses propres
        # manques sur la même période de référence.
        manquants: list[tuple[str, str]] = []
        if COMPLETER_MOIS_MANQUANTS and agregats:
            # Référentiel : tous les mois pointés, plus les trous de calendrier
            # entre le plus ancien et le plus récent. Deux manques distincts
            # sont ainsi couverts — un mois où *aucune* facture n'est arrivée,
            # et un mois où telle administration n'a rien envoyé alors qu'une
            # autre a été pointée.
            # Seules les périodes réellement interprétables font référence : une
            # période illisible (« - ») reprocherait à toutes les autres
            # administrations l'absence d'une facture pour un mois inexistant.
            reference = {
                periode for _, periode in agregats if periode_vers_cle(periode) is not None
            }
            reference.update(mois_manquants(reference))
            for administration in {adm for adm, _ in agregats}:
                pointees = {p for adm, p in agregats if adm == administration}
                manquants.extend(
                    (administration, periode)
                    for periode in reference
                    if periode not in pointees
                )

        lignes_resume: list[tuple[str, str, str, dict[str, float]]] = [
            (
                administration,
                periode,
                ETAT_INCOMPLET if agregat["non_chiffre"] else ETAT_RECUE,
                agregat,
            )
            for (administration, periode), agregat in agregats.items()
        ]
        vide = dict.fromkeys(
            (
                "lignes",
                "consommation",
                "ht",
                "tva",
                "ttc",
                "a_verifier",
                "illisible",
                "non_chiffre",
            ),
            0.0,
        )
        lignes_resume.extend(
            (administration, periode, ETAT_MANQUANTE, vide)
            for administration, periode in sorted(set(manquants))
        )

        def _ordre(entree: tuple[str, str, str, dict[str, float]]):
            administration, periode, _, _ = entree
            cle = periode_vers_cle(periode)
            return (administration, cle or (9999, 99), periode)

        for administration, periode, etat, agregat in sorted(lignes_resume, key=_ordre):
            feuille.append(
                [
                    administration,
                    periode,
                    etat,
                    int(agregat["lignes"]),
                    agregat["consommation"],
                    agregat["ht"],
                    agregat["tva"],
                    agregat["ttc"],
                    int(agregat["a_verifier"]),
                    int(agregat["illisible"]),
                ]
            )
            fond = None
            if etat == ETAT_MANQUANTE:
                fond = _FOND_MOIS_MANQUANT
            elif etat == ETAT_INCOMPLET:
                fond = _FOND_INCOMPLET
            if fond is not None:
                for index in range(1, len(_ENTETES_RESUME) + 1):
                    feuille.cell(row=feuille.max_row, column=index).fill = fond

        if agregats:
            total_incomplet = any(a["non_chiffre"] for a in agregats.values())
            feuille.append(
                [
                    "TOTAL GÉNÉRAL",
                    "",
                    "Total incomplet" if total_incomplet else "",
                    int(sum(a["lignes"] for a in agregats.values())),
                    sum(a["consommation"] for a in agregats.values()),
                    sum(a["ht"] for a in agregats.values()),
                    sum(a["tva"] for a in agregats.values()),
                    sum(a["ttc"] for a in agregats.values()),
                    int(sum(a["a_verifier"] for a in agregats.values())),
                    int(sum(a["illisible"] for a in agregats.values())),
                ]
            )
            for index in range(1, len(_ENTETES_RESUME) + 1):
                cellule = feuille.cell(row=feuille.max_row, column=index)
                cellule.font = Font(bold=True)
                if total_incomplet:
                    cellule.fill = _FOND_INCOMPLET

        # Une facture déposée mais non lue n'a produit aucune ligne : sans ce
        # renvoi, elle n'existe que dans `data/errors/` et le Résumé laisse
        # croire que tout ce qui a été déposé a été pointé.
        en_echec = 0
        if FEUILLE_ECHECS in classeur.sheetnames:
            en_echec = max(classeur[FEUILLE_ECHECS].max_row - 1, 0)
        if en_echec:
            feuille.append(
                [
                    "FACTURES EN ÉCHEC",
                    f"voir la feuille « {FEUILLE_ECHECS} »",
                    f"{en_echec} fichier(s) à retraiter",
                ]
            )
            for index in range(1, len(_ENTETES_RESUME) + 1):
                cellule = feuille.cell(row=feuille.max_row, column=index)
                cellule.font = Font(bold=True)
                cellule.fill = _FOND_ECHEC

        for ligne_index in range(2, feuille.max_row + 1):
            for colonne_index in range(5, 9):
                feuille.cell(row=ligne_index, column=colonne_index).number_format = _FORMAT_NOMBRE

    # -- API publique ------------------------------------------------------ #

    def fichier_deja_traite(self, empreinte: str) -> Optional[str]:
        """Retourne le nom du fichier déjà journalisé pour cette empreinte."""
        if not empreinte or not self.chemin.exists():
            return None
        try:
            classeur = load_workbook(self.chemin, read_only=True, data_only=True)
        except Exception as exc:
            logger.warning("Journal illisible (%s) : contrôle de doublon ignoré.", exc)
            return None
        try:
            if FEUILLE_JOURNAL not in classeur.sheetnames:
                return None
            for valeurs in classeur[FEUILLE_JOURNAL].iter_rows(min_row=2, values_only=True):
                if len(valeurs) >= 3 and valeurs[2] == empreinte:
                    return str(valeurs[1])
        finally:
            classeur.close()
        return None

    def factures_presentes(self) -> set[tuple[str, str]]:
        """Couples (compte client, période) déjà écrits dans le classeur.

        Lus dans les feuilles de données, qui font foi : le Journal peut avoir
        été purgé, les lignes réellement présentes non.
        """
        if not self.chemin.exists():
            return set()
        try:
            classeur = load_workbook(self.chemin, read_only=True, data_only=True)
        except Exception as exc:
            logger.warning("Classeur illisible (%s) : contrôle de doublon ignoré.", exc)
            return set()

        position_compte = COLONNES.index(CLE_FACTURE[0])
        position_periode = COLONNES.index(CLE_FACTURE[1])
        presentes: set[tuple[str, str]] = set()
        try:
            for onglet in self._feuilles_donnees(classeur):
                for valeurs in onglet.iter_rows(min_row=2, values_only=True):
                    if not valeurs or len(valeurs) <= position_periode:
                        continue
                    compte = str(valeurs[position_compte] or "").strip()
                    periode = str(valeurs[position_periode] or "").strip()
                    if compte and periode:
                        presentes.add((compte, periode))
        finally:
            classeur.close()
        return presentes

    def facture_deja_presente(
        self, lignes: Iterable[LigneFacture]
    ) -> list[tuple[str, str]]:
        """Identités de facture déjà présentes dans le classeur.

        Complète le contrôle par empreinte SHA-256, qui ne détecte que le même
        fichier : un re-scan, un recadrage ou un simple changement de nom
        produisent un fichier différent pour la **même** facture.
        """
        presentes = self.factures_presentes()
        return sorted(identites_facture(lignes) & presentes)

    def _purger_echecs(self, classeur: Workbook, empreinte: str, fichier: str) -> int:
        """Retire de la feuille des échecs les tentatives désormais abouties.

        La feuille des échecs est une liste de travail — ce qui reste à
        retraiter — et non un historique : une facture réintégrée avec succès
        n'a plus rien à y faire. Le rapprochement se fait sur l'empreinte quand
        elle est connue, sinon sur le nom du fichier.
        """
        if FEUILLE_ECHECS not in classeur.sheetnames:
            return 0
        feuille = classeur[FEUILLE_ECHECS]
        retirees = 0
        for index in range(feuille.max_row, 1, -1):
            ligne_empreinte = str(feuille.cell(row=index, column=3).value or "")
            ligne_fichier = str(feuille.cell(row=index, column=2).value or "")
            correspond = (
                (empreinte and ligne_empreinte == empreinte)
                or (not empreinte and ligne_fichier == fichier)
            )
            if correspond:
                feuille.delete_rows(index)
                retirees += 1
        if retirees:
            logger.info(
                "%d tentative(s) en échec retirée(s) : « %s » a finalement été intégrée.",
                retirees,
                fichier,
            )
        return retirees

    def enregistrer_echec(
        self,
        fichier: str,
        motif: str,
        empreinte: str = "",
        utilisateur: str = "",
        pages: int = 0,
    ) -> bool:
        """Inscrit au classeur une facture déposée dont la lecture a échoué.

        Une facture présentée est valable : même illisible, elle doit rester
        visible dans le pointage, sans quoi elle n'existe que dans
        `data/errors/` et rien n'indique qu'elle attend d'être retraitée.

        L'empreinte n'est **pas** portée au Journal : celui-ci sert au contrôle
        de doublon, et y inscrire un échec ferait refuser le même fichier lors
        de la nouvelle tentative — exactement l'inverse du but recherché.

        Ne lève jamais : échouer à enregistrer un échec ne doit pas masquer
        l'erreur d'origine. Retourne `True` si la ligne a bien été écrite.
        """
        try:
            with _verrou_fichier():
                classeur = self._charger()
                feuille = self._feuille(classeur, FEUILLE_ECHECS)
                feuille.append(
                    [
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        fichier,
                        empreinte,
                        utilisateur or "-",
                        pages,
                        " ".join(str(motif).split()),
                    ]
                )
                for index in range(1, len(_ENTETES_ECHECS) + 1):
                    feuille.cell(row=feuille.max_row, column=index).fill = _FOND_ECHEC
                self._reconstruire_resume(classeur)
                self._enregistrer(classeur)
        except Exception as exc:
            logger.error(
                "Impossible d'inscrire l'échec de « %s » au classeur (%s) : "
                "la facture reste tracée dans data/errors/ et son rapport JSON.",
                fichier,
                exc,
            )
            return False
        logger.info("Facture « %s » inscrite comme à retraiter dans le classeur.", fichier)
        return True

    def ajouter_lignes(
        self,
        lignes: list[LigneFacture],
        fichier: str,
        empreinte: str = "",
        utilisateur: str = "",
        pages: int = 0,
        confiance: float = 0.0,
    ) -> int:
        """Ajoute les lignes au classeur et retourne le nombre de lignes écrites.

        L'opération est atomique : soit toutes les lignes (plus les feuilles
        Anomalies, Journal et Résumé) sont écrites, soit le classeur reste
        exactement dans son état précédent.
        """
        if not lignes:
            return 0

        with _verrou_fichier():
            classeur = self._charger()
            for ligne in lignes:
                # Une ligne sans numéro de compte est nulle : elle est rangée à
                # part, hors des feuilles par ministère, donc hors de tous les
                # totaux — mais jamais supprimée.
                if ligne.statut == MARQUEUR_ECARTEE:
                    cible = FEUILLE_ECARTEES
                else:
                    cible = nom_feuille(FEUILLE_UNIQUE or ligne.administration)
                self._ecrire_ligne(self._feuille(classeur, cible), ligne, fichier)

            self._ajouter_anomalies(classeur, lignes, fichier)
            self._journaliser(classeur, fichier, empreinte, utilisateur, pages, lignes, confiance)
            # La facture est intégrée : elle n'est plus à retraiter.
            self._purger_echecs(classeur, empreinte, fichier)
            self._reconstruire_resume(classeur)
            self._enregistrer(classeur)

        logger.info("%d ligne(s) ajoutée(s) au classeur %s", len(lignes), self.chemin.name)
        return len(lignes)

    def statistiques(self) -> dict[str, Any]:
        """Chiffres clés du classeur, pour l'écran de supervision."""
        if not self.chemin.exists():
            return {
                "existe": False,
                "chemin": str(self.chemin),
                "feuilles": [],
                "lignes": 0,
                "factures_traitees": 0,
            }
        classeur = load_workbook(self.chemin, read_only=True, data_only=True)
        try:
            feuilles = []
            total = 0
            for onglet in self._feuilles_donnees(classeur):
                nb = max(onglet.max_row - 1, 0)
                total += nb
                feuilles.append({"nom": onglet.title, "lignes": nb})
            factures = 0
            if FEUILLE_JOURNAL in classeur.sheetnames:
                factures = max(classeur[FEUILLE_JOURNAL].max_row - 1, 0)
            return {
                "existe": True,
                "chemin": str(self.chemin),
                "taille_octets": self.chemin.stat().st_size,
                "feuilles": feuilles,
                "lignes": total,
                "factures_traitees": factures,
            }
        finally:
            classeur.close()
