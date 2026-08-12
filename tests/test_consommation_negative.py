"""Une consommation négative ne fabrique jamais de montant.

Un relevé incohérent — index inversés, compteur remplacé, chiffre mal lu —
produit une différence d'index négative. En tirer une consommation puis des
montants revenait à inventer une dette négative : la ligne était bien marquée
`À_VÉRIFIER`, mais ses montants entraient quand même dans les totaux et les
amputaient d'autant.

La règle « une facture présentée est valable » interdit de retirer la ligne du
total. La correction est donc en amont : on ne déduit rien d'une valeur
négative. La ligne reste au pointage, ses colonnes obligatoires sont
`ILLISIBLE`, et le total de l'administration se déclare incomplet.
"""

import pytest

from camwater.calculs import calculer_ligne
from camwater.config import MARQUEUR_ILLISIBLE
from camwater.excel_manager import ETAT_INCOMPLET, ExcelManager
from camwater.extraction import FactureExtraite
from camwater.models import RapportTraitement
from camwater.pipeline import construire_lignes
from camwater.validation import valider_rapport


# --------------------------------------------------------------------------- #
# Aucun montant déduit d'un écart négatif
# --------------------------------------------------------------------------- #


def test_index_inverses_ne_produisent_aucun_montant():
    """Le cas de l'audit : la ligne apportait -22 777 FCFA au total."""
    resultat = calculer_ligne(
        {"index_nouvel": "1 200", "index_ancien": "1 250", "location_compteur": "0"}
    )

    assert resultat.consommation is None
    assert resultat.montant_ht is None
    assert resultat.tva is None
    assert resultat.montant_ttc is None
    assert resultat.champs_derives == [], "rien ne doit être reconstruit"


def test_l_anomalie_nomme_les_deux_index():
    """Un contrôleur doit pouvoir corriger sans rouvrir le scan."""
    resultat = calculer_ligne({"index_nouvel": "1 200", "index_ancien": "1 250"})

    anomalie = " ".join(resultat.anomalies)
    assert "1200" in anomalie and "1250" in anomalie
    assert "-50" in anomalie
    assert "inversés" in anomalie


def test_montants_imprimes_lisibles_restent_exploites():
    """Les index sont douteux, pas la facture : ses montants font foi."""
    resultat = calculer_ligne(
        {
            "index_nouvel": "1 200",
            "index_ancien": "1 250",
            "location_compteur": "1 000",
            "montant_ht": "16 280",
            "tva": "3 134",
            "montant_ttc": "19 414",
        }
    )

    assert resultat.montant_ht == 16280
    assert resultat.montant_ttc == 19414
    assert resultat.consommation == 40, "récupérée depuis le HT imprimé"
    assert any("inversés" in a for a in resultat.anomalies)


def test_consommation_negative_lue_est_conservee_mais_sterile():
    """Le modèle transcrit, il ne corrige pas — mais on n'en déduit rien."""
    resultat = calculer_ligne({"consommation": "(50)", "location_compteur": "0"})

    assert resultat.consommation == -50, "la valeur lue est conservée telle quelle"
    assert resultat.montant_ht is None
    assert resultat.montant_ttc is None
    assert any("aucun montant n'en a été déduit" in a for a in resultat.anomalies)


def test_ht_inferieur_a_la_location_ne_deduit_pas_de_consommation():
    """L'autre chemin vers une consommation négative : par les montants."""
    resultat = calculer_ligne(
        {"consommation": "ILLISIBLE", "location_compteur": "5 000", "montant_ht": "1 000"}
    )

    assert resultat.consommation is None
    assert resultat.montant_ht == 1000, "le HT imprimé reste exploité"
    assert any("négative" in a for a in resultat.anomalies)


def test_aucun_index_reconstruit_depuis_une_consommation_negative():
    resultat = calculer_ligne({"consommation": "(50)", "index_ancien": "1 250"})

    assert resultat.index_nouvel is None
    assert "index_nouvel" not in resultat.champs_derives


@pytest.mark.parametrize(
    "brut",
    [
        {"index_nouvel": "1 250", "index_ancien": "1 200", "location_compteur": "1 000"},
        {"index_nouvel": "1 200", "index_ancien": "1 200"},          # consommation nulle
        {"consommation": "40", "location_compteur": "0"},
    ],
)
def test_les_cas_sains_sont_inchanges(brut):
    """La correction ne doit rien retirer aux relevés cohérents."""
    resultat = calculer_ligne(brut)

    assert resultat.consommation is not None and resultat.consommation >= 0
    assert resultat.montant_ht is not None
    assert resultat.montant_ttc is not None
    assert not any("inversés" in a or "négative" in a for a in resultat.anomalies)


def test_consommation_nulle_reste_calculable():
    """Zéro n'est pas négatif : un abonnement sans consommation se calcule."""
    resultat = calculer_ligne({"consommation": "0", "location_compteur": "1 000"})

    assert resultat.montant_ht == 1000
    assert resultat.tva == 193
    assert resultat.montant_ttc == 1193


# --------------------------------------------------------------------------- #
# Effet sur le pointage
# --------------------------------------------------------------------------- #


def _facture_avec_une_ligne_inversee():
    return FactureExtraite(
        entete={
            "dr": "DR CENTRE", "agence": "0231", "periode": "mars-2026",
            "compte_client": "0012345678", "nom_client": "MINSANTE", "ville": "YAOUNDE",
        },
        lignes=[
            {"code_abonnement": "PL1", "nom_abonne": "HOPITAL CENTRAL", "numero_compteur": "C1",
             "index_nouvel": "1 250", "index_ancien": "1 200", "location_compteur": "1 000"},
            {"code_abonnement": "PL2", "nom_abonne": "HOPITAL DE DISTRICT", "numero_compteur": "C2",
             "index_nouvel": "1 200", "index_ancien": "1 250", "location_compteur": "0"},
        ],
    )


@pytest.fixture
def rapport_valide():
    rapport = RapportTraitement(fichier="f.pdf")
    rapport.lignes = construire_lignes(_facture_avec_une_ligne_inversee())
    valider_rapport(rapport)
    return rapport


def test_la_ligne_douteuse_reste_au_pointage(rapport_valide):
    """Une facture présentée est valable : la ligne n'est pas retirée."""
    assert len(rapport_valide.lignes) == 2

    douteuse = rapport_valide.lignes[1]
    assert douteuse.statut == MARQUEUR_ILLISIBLE
    assert douteuse.valeurs["Nom abonné"] == "HOPITAL DE DISTRICT"
    assert douteuse.valeurs["Index nouvel"] == 1200, "les index lus restent visibles"
    assert douteuse.valeurs["Index ancien"] == 1250


def test_aucun_montant_negatif_n_est_ecrit(rapport_valide):
    for ligne in rapport_valide.lignes:
        for colonne in ("Consommation", "Montant HT", "TVA", "Montant TTC"):
            valeur = ligne.valeurs[colonne]
            if isinstance(valeur, (int, float)):
                assert valeur >= 0, f"{colonne} négatif écrit au pointage : {valeur}"


def test_le_total_du_ministere_n_est_plus_ampute(rapport_valide, tmp_path):
    """Avant correction, la ligne inversée retranchait 22 777 FCFA au total."""
    from openpyxl import load_workbook

    from camwater.config import FEUILLE_RESUME

    classeur = ExcelManager(tmp_path / "p.xlsx")
    classeur.ajouter_lignes(rapport_valide.lignes, fichier="f.pdf", empreinte="h1")

    resume = {
        str(r[0]): r
        for r in load_workbook(classeur.chemin)[FEUILLE_RESUME].iter_rows(
            min_row=2, values_only=True
        )
    }
    minsante = resume["MINSANTE"]

    assert minsante[7] == 23969, "seule la ligne saine compte"
    assert minsante[2] == ETAT_INCOMPLET, "et le total dit qu'il est partiel"
