"""Tests des trois règles métier ajoutées après la mise en service.

1. Le Résumé fait apparaître les mois sans facture, pour rendre les manques
   visibles dans le fichier général.
2. Une ligne sans numéro de compte client est nulle : écartée du pointage et de
   tous les totaux, mais conservée pour contrôle.
3. Une facture déjà présente (même compte client, même période) est refusée,
   même si le fichier scanné diffère.
"""

import pytest
from openpyxl import load_workbook

from camwater import pipeline, validation
from camwater.config import FEUILLE_ECARTEES, FEUILLE_RESUME, MARQUEUR_ECARTEE
from camwater.excel_manager import ETAT_MANQUANTE, ETAT_RECUE, ExcelManager, identites_facture
from camwater.extraction import FactureExtraite
from camwater.models import LigneFacture, RapportTraitement
from camwater.periodes import mois_manquants, periode_vers_cle, serie_de_mois
from camwater.pipeline import construire_lignes
from camwater.validation import valider_ligne, valider_rapport


# --------------------------------------------------------------------------- #
# 1. Mois manquants
# --------------------------------------------------------------------------- #


@pytest.mark.parametrize(
    "periode, attendu",
    [("mars-2026", (2026, 3)), ("janv-2025", (2025, 1)), ("déc-2026", (2026, 12))],
)
def test_periode_vers_cle(periode, attendu):
    assert periode_vers_cle(periode) == attendu


@pytest.mark.parametrize("periode", ["ILLISIBLE", "", None, "2026", "xxx-2026"])
def test_periode_illisible(periode):
    assert periode_vers_cle(periode) is None


def test_serie_franchit_le_changement_d_annee():
    assert serie_de_mois((2025, 11), (2026, 2)) == [
        "nov-2025",
        "déc-2025",
        "janv-2026",
        "févr-2026",
    ]


def test_mois_manquants():
    assert mois_manquants(["janv-2026", "avr-2026"]) == ["févr-2026", "mars-2026"]
    assert mois_manquants(["janv-2026", "févr-2026"]) == []
    assert mois_manquants(["janv-2026"]) == []          # une seule période : pas d'étendue
    assert mois_manquants(["ILLISIBLE", "janv-2026"]) == []


def _ligne(ministere="MINSANTE", periode="mars-2026", compte="0012345678", montant=39200):
    return LigneFacture(
        numero=1,
        valeurs={
            "DR": "DR CENTRE",
            "Agence": "0231",
            "Période": periode,
            "Compte client": compte,
            "Nom du client": "MINISTERE DE LA SANTE",
            "Ville": "YAOUNDE",
            "Code abonnement (PL)": "PL-4455",
            "Nom abonné": "HOPITAL CENTRAL",
            "N° compteur": "A123456",
            "Index nouvel": 1350,
            "Index ancien": 1250,
            "Consommation": 100,
            "Location compteur": 1000,
            "TVA": 7546,
            "Montant HT": montant,
            "Ministère_2026": ministere,
            "Montant TTC": montant + 7546,
        },
    )


def test_resume_signale_les_mois_sans_facture(tmp_path):
    """Janvier et mars pointés : février doit apparaître comme MANQUANTE."""
    gestionnaire = ExcelManager(tmp_path / "pointage.xlsx")
    gestionnaire.ajouter_lignes([_ligne(periode="janv-2026")], fichier="f1.pdf", empreinte="h1")
    gestionnaire.ajouter_lignes([_ligne(periode="mars-2026")], fichier="f2.pdf", empreinte="h2")

    resume = load_workbook(gestionnaire.chemin)[FEUILLE_RESUME]
    lignes = [r for r in resume.iter_rows(min_row=2, values_only=True) if r[0] != "TOTAL GÉNÉRAL"]
    par_periode = {r[1]: r for r in lignes}

    assert par_periode["janv-2026"][2] == ETAT_RECUE
    assert par_periode["mars-2026"][2] == ETAT_RECUE
    assert "févr-2026" in par_periode, "le mois sans facture doit figurer au Résumé"
    assert par_periode["févr-2026"][2] == ETAT_MANQUANTE
    assert par_periode["févr-2026"][3] == 0        # aucune ligne
    assert par_periode["févr-2026"][5] == 0        # aucun montant


def test_mois_manquant_par_administration(tmp_path):
    """Un ministère absent sur un mois couvert par un autre doit être signalé."""
    gestionnaire = ExcelManager(tmp_path / "pointage.xlsx")
    gestionnaire.ajouter_lignes(
        [_ligne("MINSANTE", "janv-2026"), _ligne("MINESEC", "janv-2026")],
        fichier="f1.pdf",
        empreinte="h1",
    )
    # En février, seule MINSANTE a été reçue.
    gestionnaire.ajouter_lignes([_ligne("MINSANTE", "févr-2026")], fichier="f2.pdf", empreinte="h2")

    resume = load_workbook(gestionnaire.chemin)[FEUILLE_RESUME]
    lignes = [r for r in resume.iter_rows(min_row=2, values_only=True) if r[0] != "TOTAL GÉNÉRAL"]
    manquantes = {(r[0], r[1]) for r in lignes if r[2] == ETAT_MANQUANTE}

    assert ("MINESEC", "févr-2026") in manquantes
    assert ("MINSANTE", "févr-2026") not in manquantes


def test_les_mois_manquants_ne_faussent_pas_le_total(tmp_path):
    gestionnaire = ExcelManager(tmp_path / "pointage.xlsx")
    gestionnaire.ajouter_lignes([_ligne(periode="janv-2026")], fichier="f1.pdf", empreinte="h1")
    gestionnaire.ajouter_lignes([_ligne(periode="mars-2026")], fichier="f2.pdf", empreinte="h2")

    resume = load_workbook(gestionnaire.chemin)[FEUILLE_RESUME]
    total = [r for r in resume.iter_rows(min_row=2, values_only=True) if r[0] == "TOTAL GÉNÉRAL"][0]
    assert total[3] == 2                 # 2 lignes réelles, les mois vides n'ajoutent rien
    assert total[5] == 2 * 39200


# --------------------------------------------------------------------------- #
# 2. Ligne sans numéro de compte
# --------------------------------------------------------------------------- #


@pytest.mark.parametrize("compte", [None, "", "   ", "ILLISIBLE"])
def test_ligne_sans_compte_est_ecartee(compte):
    ligne = valider_ligne(_ligne(compte=compte))
    assert ligne.statut == MARQUEUR_ECARTEE
    assert ligne.est_ecartee
    assert any("numéro de compte" in a for a in ligne.anomalies)


def test_ligne_ecartee_reste_ecartee():
    """Le statut ÉCARTÉE est terminal : aucun contrôle ne le requalifie."""
    ligne = valider_ligne(_ligne(compte=""))
    ligne.marquer("À_VÉRIFIER", "autre anomalie")
    assert ligne.statut == MARQUEUR_ECARTEE


def test_ligne_ecartee_exclue_des_totaux():
    rapport = RapportTraitement(fichier="f.pdf")
    rapport.lignes = [_ligne(), _ligne(compte="")]
    rapport.totaux_lus = {"HT": "39 200", "TVA": "7 546", "TTC": "46 746"}

    valider_rapport(rapport)

    # Seule la ligne valide est sommée : aucun écart de total signalé.
    assert rapport.totaux_calcules["HT"] == "39200"
    assert rapport.nb_ecartees == 1
    assert not any("Total HT" in a for a in rapport.anomalies)


def test_ligne_ecartee_rangee_dans_sa_feuille(tmp_path):
    gestionnaire = ExcelManager(tmp_path / "pointage.xlsx")
    valide, nulle = _ligne(), valider_ligne(_ligne(compte=""))
    gestionnaire.ajouter_lignes([valide, nulle], fichier="f1.pdf", empreinte="h1")

    classeur = load_workbook(gestionnaire.chemin)
    assert classeur["MINSANTE"].max_row == 2          # entête + 1 ligne valide
    assert FEUILLE_ECARTEES in classeur.sheetnames
    assert classeur[FEUILLE_ECARTEES].max_row == 2    # entête + 1 ligne écartée

    # Elle ne pèse dans aucun total du Résumé.
    resume = classeur[FEUILLE_RESUME]
    total = [r for r in resume.iter_rows(min_row=2, values_only=True) if r[0] == "TOTAL GÉNÉRAL"][0]
    assert total[3] == 1


def test_ligne_sans_compte_extraite_du_pdf_est_ecartee():
    """Cas réel : l'en-tête n'a pas de compte client lisible."""
    facture = FactureExtraite(
        entete={"agence": "0231", "periode": "MARS 2026", "nom_client": "MINSANTE"},
        lignes=[{"code_abonnement": "PL-1", "nom_abonne": "HOPITAL", "consommation": "10"}],
    )
    ligne = valider_ligne(construire_lignes(facture)[0])
    assert ligne.statut == MARQUEUR_ECARTEE


# --------------------------------------------------------------------------- #
# 3. Doublon de facture (compte client + période)
# --------------------------------------------------------------------------- #


def test_identite_facture():
    assert identites_facture([_ligne()]) == {("0012345678", "mars-2026")}
    # Une ligne écartée n'a pas d'identité exploitable.
    assert identites_facture([valider_ligne(_ligne(compte=""))]) == set()


def test_facture_deja_presente_detectee(tmp_path):
    gestionnaire = ExcelManager(tmp_path / "pointage.xlsx")
    gestionnaire.ajouter_lignes([_ligne()], fichier="scan-original.pdf", empreinte="h1")

    # Même compte, même période : c'est la même facture, même re-scannée.
    doublon = gestionnaire.facture_deja_presente([_ligne()])
    assert doublon == [("0012345678", "mars-2026")]

    # Autre période : pas un doublon.
    assert gestionnaire.facture_deja_presente([_ligne(periode="avr-2026")]) == []
    # Autre compte : pas un doublon.
    assert gestionnaire.facture_deja_presente([_ligne(compte="0099999999")]) == []


def test_classeur_absent_aucun_doublon(tmp_path):
    gestionnaire = ExcelManager(tmp_path / "inexistant.xlsx")
    assert gestionnaire.factures_presentes() == set()
    assert gestionnaire.facture_deja_presente([_ligne()]) == []


@pytest.fixture
def environnement(tmp_path, monkeypatch):
    dossiers = {}
    for nom in ("uploads", "processed", "errors", "reports", "output"):
        (tmp_path / nom).mkdir()
        dossiers[nom] = tmp_path / nom
    monkeypatch.setattr(pipeline, "UPLOAD_DIR", dossiers["uploads"])
    monkeypatch.setattr(pipeline, "PROCESSED_DIR", dossiers["processed"])
    monkeypatch.setattr(pipeline, "ERROR_DIR", dossiers["errors"])
    monkeypatch.setattr(validation, "REPORT_DIR", dossiers["reports"])
    dossiers["excel"] = ExcelManager(dossiers["output"] / "pointage.xlsx")
    return dossiers


def _facture(compte="0012345678"):
    return FactureExtraite(
        entete={
            "dr": "DR CENTRE",
            "agence": "0231",
            "periode": "MARS 2026",
            "compte_client": compte,
            "nom_client": "MINISTERE DE LA SANTE PUBLIQUE",
            "ville": "YAOUNDE",
        },
        lignes=[
            {
                "code_abonnement": "PL-4455",
                "nom_abonne": "HOPITAL CENTRAL DE YAOUNDE",
                "numero_compteur": "A123456",
                "index_nouvel": "1 350",
                "index_ancien": "1 250",
                "consommation": "100",
                "location_compteur": "1 000",
                "tva": "7 546",
                "montant_ht": "39 200",
                "montant_ttc": "46 746",
            }
        ],
        total_ht="39 200",
        total_tva="7 546",
        total_ttc="46 746",
        confiance=0.95,
        pages=1,
    )


def test_rescan_de_la_meme_facture_refuse(environnement, monkeypatch):
    """Fichier différent (donc empreinte différente) mais même facture métier."""
    monkeypatch.setattr(pipeline, "extraire_facture", lambda pages: _facture())

    original = environnement["uploads"] / "scan-original.png"
    original.write_bytes(b"\x89PNG\r\n\x1a\n" + b"A" * 100)
    premier = pipeline.traiter_fichier(original, excel=environnement["excel"])
    assert premier.succes, premier.erreur

    # Contenu volontairement différent : le contrôle SHA-256 ne peut rien voir.
    rescan = environnement["uploads"] / "scan-recadre.png"
    rescan.write_bytes(b"\x89PNG\r\n\x1a\n" + b"B" * 250)
    second = pipeline.traiter_fichier(rescan, excel=environnement["excel"])

    assert second.succes is False
    assert "déjà présente" in second.erreur
    assert "0012345678" in second.erreur and "mars-2026" in second.erreur

    # Aucune ligne dupliquée dans le classeur.
    classeur = load_workbook(environnement["excel"].chemin)
    assert classeur["MINSANTE"].max_row == 2


def test_facture_d_un_autre_compte_acceptee(environnement, monkeypatch):
    monkeypatch.setattr(pipeline, "extraire_facture", lambda pages: _facture())
    premier = environnement["uploads"] / "a.png"
    premier.write_bytes(b"\x89PNG\r\n\x1a\n" + b"A" * 100)
    pipeline.traiter_fichier(premier, excel=environnement["excel"])

    monkeypatch.setattr(pipeline, "extraire_facture", lambda pages: _facture("0099999999"))
    second = environnement["uploads"] / "b.png"
    second.write_bytes(b"\x89PNG\r\n\x1a\n" + b"B" * 100)
    rapport = pipeline.traiter_fichier(second, excel=environnement["excel"])

    assert rapport.succes, rapport.erreur
    assert load_workbook(environnement["excel"].chemin)["MINSANTE"].max_row == 3


# --------------------------------------------------------------------------- #
# Compte client au niveau de la ligne
# --------------------------------------------------------------------------- #


def _facture_multi_comptes(compte_ligne_2):
    """Facture récapitulative portant une colonne « compte » par ligne."""
    return FactureExtraite(
        entete={"agence": "0231", "periode": "MARS 2026", "compte_client": "0011111111"},
        lignes=[
            {"compte_client": "0022222222", "nom_abonne": "HOPITAL", "consommation": "10",
             "montant_ht": "3 820", "tva": "735", "montant_ttc": "4 555"},
            {"compte_client": compte_ligne_2, "nom_abonne": "LYCEE", "consommation": "20",
             "montant_ht": "7 640", "tva": "1 471", "montant_ttc": "9 111"},
        ],
    )


def test_compte_de_ligne_prime_sur_l_entete():
    lignes = construire_lignes(_facture_multi_comptes("0033333333"))
    assert lignes[0].valeurs["Compte client"] == "0022222222"
    assert lignes[1].valeurs["Compte client"] == "0033333333"


def test_compte_de_ligne_vide_reprend_celui_de_l_entete():
    """Colonne absente du tableau : l'en-tête fait foi, la ligne reste valide."""
    ligne = valider_ligne(construire_lignes(_facture_multi_comptes(""))[1])
    assert ligne.valeurs["Compte client"] == "0011111111"
    assert ligne.statut != MARQUEUR_ECARTEE


def test_compte_de_ligne_illisible_ecarte_la_seule_ligne_concernee():
    """Colonne présente mais indéchiffrable : cette ligne seule est écartée."""
    lignes = [valider_ligne(l) for l in construire_lignes(_facture_multi_comptes("ILLISIBLE"))]
    assert lignes[0].statut != MARQUEUR_ECARTEE      # l'autre ligne reste intègre
    assert lignes[1].statut == MARQUEUR_ECARTEE
