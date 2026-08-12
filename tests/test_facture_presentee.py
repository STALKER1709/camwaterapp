"""Règle métier : « tant qu'une facture est présentée, elle est valable ».

Une facture déposée ne doit jamais disparaître du pointage. Deux conséquences
distinctes, testées séparément ici :

1. une facture **non lisible** reste inscrite au classeur comme restant à
   retraiter — sans quoi elle n'existerait que dans `data/errors/` ;
2. un total calculé à partir de chiffres partiellement illisibles est un
   **minorant**, et doit se présenter comme tel plutôt que comme un total
   définitif.

Un doublon échappe à la première règle : la facture est déjà pointée, elle
n'est pas « en attente de retraitement ».
"""

import pytest
from openpyxl import load_workbook

from camwater.config import (
    FEUILLE_ECHECS,
    FEUILLE_RESUME,
    MARQUEUR_A_VERIFIER,
    MARQUEUR_ILLISIBLE,
)
from camwater.excel_manager import ETAT_INCOMPLET, ETAT_RECUE, ExcelManager
from camwater.models import STATUT_OK, LigneFacture


@pytest.fixture
def classeur(tmp_path):
    return ExcelManager(tmp_path / "pointage.xlsx")


def _ligne(ministere="MINSANTE", statut=STATUT_OK, ht=15280, tva=2941, ttc=18221,
           conso=40, periode="mars-2026", compte="123456"):
    return LigneFacture(statut=statut, valeurs={
        "DR": "DR CENTRE", "Agence": "0231", "Période": periode,
        "Compte client": compte, "Nom du client": "MINSANTE", "Ville": "YAOUNDE",
        "Code abonnement (PL)": "PL1", "Nom abonné": "HOPITAL CENTRAL",
        "N° compteur": "C1", "Index nouvel": 100, "Index ancien": 60,
        "Consommation": conso, "Location compteur": 0, "TVA": tva,
        "Montant HT": ht, "Ministère_2026": ministere, "Montant TTC": ttc})


def _resume(chemin):
    """Le Résumé sous forme de tuples (administration, période, état)."""
    feuille = load_workbook(chemin)[FEUILLE_RESUME]
    return [
        (str(r[0]), str(r[1]), str(r[2]))
        for r in feuille.iter_rows(min_row=2, values_only=True)
    ]


# --------------------------------------------------------------------------- #
# 1. Une facture illisible reste visible
# --------------------------------------------------------------------------- #


def test_facture_en_echec_inscrite_au_classeur(classeur):
    assert classeur.enregistrer_echec(
        fichier="MinSante.pdf",
        motif="Le crédit du compte API Anthropic est épuisé.",
        empreinte="abc123",
        utilisateur="poste-3",
        pages=2,
    )

    lignes = list(load_workbook(classeur.chemin)[FEUILLE_ECHECS].iter_rows(
        min_row=2, values_only=True))
    assert len(lignes) == 1
    _, fichier, empreinte, utilisateur, pages, motif = lignes[0]
    assert (fichier, empreinte, utilisateur, pages) == ("MinSante.pdf", "abc123", "poste-3", 2)
    assert "crédit" in motif


def test_le_resume_renvoie_vers_les_factures_en_echec(classeur):
    classeur.enregistrer_echec(fichier="a.pdf", motif="illisible", empreinte="a")
    classeur.enregistrer_echec(fichier="b.pdf", motif="illisible", empreinte="b")

    renvois = [r for r in _resume(classeur.chemin) if r[0] == "FACTURES EN ÉCHEC"]
    assert len(renvois) == 1
    assert "2 fichier(s) à retraiter" in renvois[0][2]


def test_l_echec_n_empeche_pas_une_nouvelle_tentative(classeur):
    """L'empreinte d'un échec ne doit pas atterrir au Journal.

    Le Journal sert au contrôle de doublon : y inscrire un échec ferait
    refuser le même fichier lors de la nouvelle tentative — exactement
    l'inverse du but recherché.
    """
    classeur.enregistrer_echec(fichier="MinSante.pdf", motif="crédit épuisé", empreinte="abc123")

    assert classeur.fichier_deja_traite("abc123") is None


def test_la_facture_reintegree_quitte_la_liste(classeur):
    classeur.enregistrer_echec(fichier="MinSante.pdf", motif="crédit épuisé", empreinte="abc123")
    classeur.ajouter_lignes([_ligne()], fichier="MinSante.pdf", empreinte="abc123")

    feuille = load_workbook(classeur.chemin)[FEUILLE_ECHECS]
    assert feuille.max_row == 1, "la facture intégrée ne reste pas « à retraiter »"
    assert not [r for r in _resume(classeur.chemin) if r[0] == "FACTURES EN ÉCHEC"]


def test_seule_la_facture_concernee_quitte_la_liste(classeur):
    classeur.enregistrer_echec(fichier="a.pdf", motif="illisible", empreinte="aaa")
    classeur.enregistrer_echec(fichier="b.pdf", motif="illisible", empreinte="bbb")
    classeur.ajouter_lignes([_ligne()], fichier="a.pdf", empreinte="aaa")

    restants = [
        r[1] for r in load_workbook(classeur.chemin)[FEUILLE_ECHECS].iter_rows(
            min_row=2, values_only=True)
    ]
    assert restants == ["b.pdf"]


def test_echec_d_ecriture_ne_masque_pas_l_erreur_d_origine(classeur, monkeypatch):
    """Échouer à enregistrer un échec ne doit jamais lever."""
    def refus(_classeur):
        raise OSError("classeur verrouillé par Excel")

    monkeypatch.setattr(ExcelManager, "_enregistrer", refus)
    assert classeur.enregistrer_echec(fichier="x.pdf", motif="illisible") is False


# --------------------------------------------------------------------------- #
# 2. Un total bâti sur des chiffres illisibles est un minorant
# --------------------------------------------------------------------------- #


def test_total_complet_non_marque(classeur):
    classeur.ajouter_lignes([_ligne(), _ligne(compte="2")], fichier="f.pdf")

    etats = {r[2] for r in _resume(classeur.chemin) if r[0] == "MINSANTE"}
    assert etats == {ETAT_RECUE}


def test_un_montant_illisible_marque_le_total_incomplet(classeur):
    classeur.ajouter_lignes(
        [
            _ligne(),
            _ligne(compte="2", statut=MARQUEUR_ILLISIBLE, ht=MARQUEUR_ILLISIBLE,
                   tva=MARQUEUR_ILLISIBLE, ttc=MARQUEUR_ILLISIBLE, conso=MARQUEUR_ILLISIBLE),
        ],
        fichier="f.pdf",
    )

    ligne = next(r for r in _resume(classeur.chemin) if r[0] == "MINSANTE")
    assert ligne[2] == ETAT_INCOMPLET
    assert "incomplet" in ligne[2].lower()


def test_le_total_general_est_marque_aussi(classeur):
    classeur.ajouter_lignes(
        [_ligne(), _ligne(ministere="MINESEC", compte="2", ttc=MARQUEUR_ILLISIBLE)],
        fichier="f.pdf",
    )

    total = next(r for r in _resume(classeur.chemin) if r[0] == "TOTAL GÉNÉRAL")
    assert total[2] == "Total incomplet"


def test_seule_l_administration_touchee_est_marquee(classeur):
    classeur.ajouter_lignes(
        [_ligne(), _ligne(ministere="MINESEC", compte="2", ttc=MARQUEUR_ILLISIBLE)],
        fichier="f.pdf",
    )
    etats = {r[0]: r[2] for r in _resume(classeur.chemin)}

    assert etats["MINSANTE"] == ETAT_RECUE
    assert etats["MINESEC"] == ETAT_INCOMPLET


def test_une_ligne_a_verifier_chiffree_ne_marque_rien(classeur):
    """« À vérifier » n'est pas « illisible » : les chiffres sont là et comptent."""
    classeur.ajouter_lignes([_ligne(statut=MARQUEUR_A_VERIFIER)], fichier="f.pdf")

    ligne = next(r for r in _resume(classeur.chemin) if r[0] == "MINSANTE")
    assert ligne[2] == ETAT_RECUE


# --------------------------------------------------------------------------- #
# 3. Régressions relevées à l'audit, dans la même fonction
# --------------------------------------------------------------------------- #


def test_periode_illisible_ne_fabrique_pas_de_mois_manquant(classeur):
    """Régression : « - » entrait dans le référentiel des périodes."""
    classeur.ajouter_lignes(
        [
            _ligne(),
            _ligne(ministere="MINESEC", compte="2"),
            _ligne(compte="3", periode=None),          # période illisible
        ],
        fichier="f.pdf",
    )

    manquants = [r for r in _resume(classeur.chemin) if r[2] == "MANQUANTE"]
    assert manquants == [], f"lignes MANQUANTE parasites : {manquants}"


def test_feuille_trop_etroite_n_interrompt_pas_le_pointage(classeur):
    """Régression : IndexError qui bloquait toute écriture ultérieure."""
    classeur.ajouter_lignes([_ligne()], fichier="f.pdf")

    livre = load_workbook(classeur.chemin)
    livre.create_sheet("Notes").append(["DR", "Agence", "Période"])
    livre.save(classeur.chemin)

    classeur.ajouter_lignes([_ligne(compte="2")], fichier="g.pdf")   # ne doit pas lever

    ligne = next(r for r in _resume(classeur.chemin) if r[0] == "MINSANTE")
    assert ligne[2] == ETAT_RECUE
