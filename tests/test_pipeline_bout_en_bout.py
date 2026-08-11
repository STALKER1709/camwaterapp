"""Test de bout en bout du pipeline, avec la lecture visuelle simulée.

L'appel au modèle est remplacé par une extraction figée : on vérifie donc toute
la chaîne (conversion, calculs, mapping, validation, Excel, archivage, rapport)
sans dépendre du réseau ni d'une clé API.
"""

import json

import pytest
from openpyxl import load_workbook

from camwater import pipeline, validation
from camwater.excel_manager import ExcelManager
from camwater.extraction import ExtractionError, FactureExtraite


@pytest.fixture
def environnement(tmp_path, monkeypatch):
    """Redirige tous les dossiers de travail vers un répertoire temporaire."""
    dossiers = {}
    for nom in ("uploads", "processed", "errors", "reports", "output"):
        chemin = tmp_path / nom
        chemin.mkdir()
        dossiers[nom] = chemin

    monkeypatch.setattr(pipeline, "UPLOAD_DIR", dossiers["uploads"])
    monkeypatch.setattr(pipeline, "PROCESSED_DIR", dossiers["processed"])
    monkeypatch.setattr(pipeline, "ERROR_DIR", dossiers["errors"])
    monkeypatch.setattr(validation, "REPORT_DIR", dossiers["reports"])
    dossiers["excel"] = ExcelManager(dossiers["output"] / "CAMWATER_Pointage_General.xlsx")
    return dossiers


def _facture_simulee() -> FactureExtraite:
    return FactureExtraite(
        entete={
            "dr": "DR CENTRE",
            "agence": "0231",
            "periode": "MARS 2026",
            "compte_client": "0012345678",
            "nom_client": "MINISTERE DE LA SANTE PUBLIQUE",
            "ville": "YAOUNDE",
        },
        lignes=[
            {
                "code_abonnement": "PL-4455",
                "nom_abonne": "HOPITAL CENTRAL DE YAOUNDE",
                "numero_compteur": "A123456",
                "index_nouvel": "1 350",
                "index_ancien": "1 250",
                "consommation": "100",
                "location_compteur": "1 000",
                "tva": "7 546",
                "montant_ht": "39 200",
                "montant_ttc": "46 746",
            },
            {
                "code_abonnement": "PL-4456",
                "nom_abonne": "CENTRE DE SANTE INTEGRE DE NKOLBISSON",
                "numero_compteur": "B778899",
                "index_nouvel": "560",
                "index_ancien": "500",
                "consommation": "60",
                "location_compteur": "0",
                "tva": "4 412",
                "montant_ht": "22 920",
                "montant_ttc": "27 332",
            },
        ],
        total_ht="62 120",
        total_tva="11 958",
        total_ttc="74 078",
        confiance=0.96,
        pages=1,
    )


def _facture_png(dossier, nom="facture-test.png"):
    chemin = dossier / nom
    chemin.write_bytes(b"\x89PNG\r\n\x1a\n" + b"0" * 128)  # contenu factice : lecture simulée
    return chemin


def test_traitement_reussi(environnement, monkeypatch):
    monkeypatch.setattr(pipeline, "extraire_facture", lambda pages: _facture_simulee())
    source = _facture_png(environnement["uploads"])

    rapport = pipeline.traiter_fichier(source, excel=environnement["excel"], utilisateur="poste-1")

    assert rapport.succes, rapport.erreur
    assert rapport.lignes_ecrites == 2
    assert rapport.nb_ok == 2
    assert rapport.nb_a_verifier == 0
    assert rapport.totaux_calcules["TTC"] == "74078"

    # Le fichier a été archivé, pas laissé dans uploads/.
    assert not source.exists()
    assert (environnement["processed"] / "facture-test.png").exists()

    # Le rapport JSON est écrit et relisible.
    rapport_json = json.loads((environnement["reports"] / "facture-test.rapport.json").read_text("utf-8"))
    assert rapport_json["succes"] is True
    assert rapport_json["statistiques"]["lignes_ecrites"] == 2

    # Les lignes sont dans la bonne feuille du classeur.
    classeur = load_workbook(environnement["excel"].chemin)
    assert classeur["MINSANTE"].max_row == 3  # entête + 2 lignes
    assert classeur["MINSANTE"].cell(row=2, column=3).value == "mars-2026"


def test_doublon_refuse(environnement, monkeypatch):
    monkeypatch.setattr(pipeline, "extraire_facture", lambda pages: _facture_simulee())

    premier = _facture_png(environnement["uploads"], "facture-a.png")
    pipeline.traiter_fichier(premier, excel=environnement["excel"])

    # Même contenu, nom différent : l'empreinte SHA-256 est identique.
    second = _facture_png(environnement["uploads"], "facture-b.png")
    rapport = pipeline.traiter_fichier(second, excel=environnement["excel"])

    assert rapport.succes is False
    assert "déjà intégrée" in rapport.erreur
    assert (environnement["errors"] / "facture-b.png").exists()

    classeur = load_workbook(environnement["excel"].chemin)
    assert classeur["MINSANTE"].max_row == 3  # inchangé : aucune ligne en double


def test_echec_de_lecture_archive_en_erreur(environnement, monkeypatch):
    def lecture_impossible(pages):
        raise ExtractionError("Aucune ligne de facturation lisible")

    monkeypatch.setattr(pipeline, "extraire_facture", lecture_impossible)
    source = _facture_png(environnement["uploads"], "illisible.png")

    rapport = pipeline.traiter_fichier(source, excel=environnement["excel"])

    assert rapport.succes is False
    assert "lisible" in rapport.erreur
    assert (environnement["errors"] / "illisible.png").exists()
    assert (environnement["reports"] / "illisible.rapport.json").exists()
    # Aucun classeur créé : rien n'a été écrit.
    assert not environnement["excel"].chemin.exists()


def test_incoherence_de_totaux_signalee_sans_blocage(environnement, monkeypatch):
    facture = _facture_simulee()
    facture.total_ttc = "99 999"
    monkeypatch.setattr(pipeline, "extraire_facture", lambda pages: facture)
    source = _facture_png(environnement["uploads"], "ecart.png")

    rapport = pipeline.traiter_fichier(source, excel=environnement["excel"])

    # Les lignes sont écrites (rien n'est perdu) mais toutes marquées à vérifier.
    assert rapport.succes is True
    assert rapport.lignes_ecrites == 2
    assert rapport.nb_a_verifier == 2
    assert any("Total TTC" in anomalie for anomalie in rapport.anomalies)
