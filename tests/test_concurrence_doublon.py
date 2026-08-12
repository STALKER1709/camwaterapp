"""Le contrôle anti-doublon doit résister à deux dépôts simultanés.

L'application est faite pour que plusieurs postes déposent en parallèle. Tant
que le contrôle précédait la prise du verrou, deux postes pouvaient constater
tous deux l'absence de la facture, puis l'écrire tous deux : la règle « éviter
les doublons » sautait précisément dans le scénario qu'elle vise.

Ces tests exercent la concurrence pour de bon — threads et processus séparés —
plutôt que de vérifier l'ordre des lignes de code.
"""

import multiprocessing
import sys
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import pytest
from openpyxl import load_workbook

from camwater.excel_manager import DoublonError, ExcelManager
from camwater.models import LigneFacture


def _ligne(compte="0012345678", periode="mars-2026"):
    return LigneFacture(valeurs={
        "DR": "DR CENTRE", "Agence": "0231", "Période": periode,
        "Compte client": compte, "Nom du client": "MINSANTE", "Ville": "YAOUNDE",
        "Code abonnement (PL)": "PL1", "Nom abonné": "HOPITAL CENTRAL",
        "N° compteur": "C1", "Index nouvel": 100, "Index ancien": 60,
        "Consommation": 40, "Location compteur": 0, "TVA": 2941,
        "Montant HT": 15280, "Ministère_2026": "MINSANTE", "Montant TTC": 18221})


def _lignes_ecrites(chemin: Path) -> int:
    return max(load_workbook(chemin)["MINSANTE"].max_row - 1, 0)


# --------------------------------------------------------------------------- #
# Séquentiel : la base
# --------------------------------------------------------------------------- #


def test_seconde_ecriture_refusee(tmp_path):
    classeur = ExcelManager(tmp_path / "p.xlsx")
    classeur.ajouter_lignes([_ligne()], fichier="a.pdf", empreinte="aaa")

    with pytest.raises(DoublonError, match="déjà présente"):
        classeur.ajouter_lignes([_ligne()], fichier="b.pdf", empreinte="bbb")

    assert _lignes_ecrites(classeur.chemin) == 1


def test_meme_fichier_refuse_sur_l_empreinte(tmp_path):
    classeur = ExcelManager(tmp_path / "p.xlsx")
    classeur.ajouter_lignes([_ligne()], fichier="a.pdf", empreinte="aaa")

    with pytest.raises(DoublonError, match="déjà intégrée"):
        classeur.ajouter_lignes([_ligne(compte="999")], fichier="a.pdf", empreinte="aaa")


def test_facture_distincte_acceptee(tmp_path):
    classeur = ExcelManager(tmp_path / "p.xlsx")
    classeur.ajouter_lignes([_ligne()], fichier="a.pdf", empreinte="aaa")
    classeur.ajouter_lignes([_ligne(periode="avr-2026")], fichier="b.pdf", empreinte="bbb")

    assert _lignes_ecrites(classeur.chemin) == 2


# --------------------------------------------------------------------------- #
# Concurrence réelle
# --------------------------------------------------------------------------- #


def test_dix_threads_simultanes_n_ecrivent_qu_une_fois(tmp_path):
    """Dix postes déposent la même facture en même temps."""
    classeur = ExcelManager(tmp_path / "p.xlsx")

    def deposer(numero):
        try:
            classeur.ajouter_lignes(
                [_ligne()], fichier=f"poste-{numero}.pdf", empreinte=f"h{numero}"
            )
            return "écrit"
        except DoublonError:
            return "refusé"

    with ThreadPoolExecutor(max_workers=10) as executeur:
        resultats = list(executeur.map(deposer, range(10)))

    assert resultats.count("écrit") == 1, f"une seule écriture attendue : {resultats}"
    assert resultats.count("refusé") == 9
    assert _lignes_ecrites(classeur.chemin) == 1


def test_factures_distinctes_en_parallele_passent_toutes(tmp_path):
    """Le verrou sérialise sans jamais rejeter une facture légitime."""
    classeur = ExcelManager(tmp_path / "p.xlsx")

    def deposer(numero):
        classeur.ajouter_lignes(
            [_ligne(compte=f"compte-{numero}")],
            fichier=f"f{numero}.pdf",
            empreinte=f"h{numero}",
        )

    with ThreadPoolExecutor(max_workers=8) as executeur:
        list(executeur.map(deposer, range(8)))

    assert _lignes_ecrites(classeur.chemin) == 8


# --------------------------------------------------------------------------- #
# Processus distincts : le cas réel du multi-poste
# --------------------------------------------------------------------------- #


def _deposer_depuis_un_autre_processus(arguments):
    """Exécuté dans un processus fils : import complet, verrou `filelock`."""
    chemin, numero = arguments
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from camwater.excel_manager import DoublonError as Doublon
    from camwater.excel_manager import ExcelManager as Gestionnaire
    from camwater.models import LigneFacture as Ligne

    ligne = Ligne(valeurs={
        "DR": "DR CENTRE", "Agence": "0231", "Période": "mars-2026",
        "Compte client": "0012345678", "Nom du client": "MINSANTE", "Ville": "YAOUNDE",
        "Code abonnement (PL)": "PL1", "Nom abonné": "HOPITAL CENTRAL",
        "N° compteur": "C1", "Index nouvel": 100, "Index ancien": 60,
        "Consommation": 40, "Location compteur": 0, "TVA": 2941,
        "Montant HT": 15280, "Ministère_2026": "MINSANTE", "Montant TTC": 18221})
    try:
        Gestionnaire(Path(chemin)).ajouter_lignes(
            [ligne], fichier=f"poste-{numero}.pdf", empreinte=f"h{numero}"
        )
        return "écrit"
    except Doublon:
        return "refusé"


def test_processus_concurrents_n_ecrivent_qu_une_fois(tmp_path):
    """Le verrou `filelock` doit tenir entre processus, pas seulement threads."""
    chemin = str(tmp_path / "p.xlsx")
    contexte = multiprocessing.get_context("spawn")

    with contexte.Pool(4) as reservoir:
        resultats = reservoir.map(
            _deposer_depuis_un_autre_processus, [(chemin, n) for n in range(4)]
        )

    assert resultats.count("écrit") == 1, f"une seule écriture attendue : {resultats}"
    assert _lignes_ecrites(Path(chemin)) == 1
