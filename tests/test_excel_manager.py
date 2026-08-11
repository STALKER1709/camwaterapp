"""Tests d'écriture du classeur Excel général."""

from openpyxl import load_workbook

from camwater.config import (
    COLONNES,
    FEUILLE_ANOMALIES,
    FEUILLE_JOURNAL,
    FEUILLE_RESUME,
    MARQUEUR_A_VERIFIER,
)
from camwater.excel_manager import COLONNES_TECHNIQUES, ExcelManager, nom_feuille
from camwater.models import LigneFacture


def _ligne(ministere="MINSANTE", montant_ht=39200, statut="OK", periode="mars-2026"):
    return LigneFacture(
        numero=1,
        statut=statut,
        valeurs={
            "DR": "DR CENTRE",
            "Agence": "0231",
            "Période": periode,
            "Compte client": "0012345678",
            "Nom du client": "MINISTERE DE LA SANTE",
            "Ville": "YAOUNDE",
            "Code abonnement (PL)": "PL-4455",
            "Nom abonné": "HOPITAL CENTRAL",
            "N° compteur": "A123456",
            "Index nouvel": 1350,
            "Index ancien": 1250,
            "Consommation": 100,
            "Location compteur": 1000,
            "TVA": 7546,
            "Montant HT": montant_ht,
            "Ministère_2026": ministere,
            "Montant TTC": montant_ht + 7546,
        },
    )


def test_nom_feuille_assaini():
    assert nom_feuille("MIN/SANTE") == "MIN-SANTE"
    assert nom_feuille("") == MARQUEUR_A_VERIFIER
    assert len(nom_feuille("A" * 50)) == 31


def test_creation_du_classeur_et_des_feuilles(tmp_path):
    gestionnaire = ExcelManager(tmp_path / "pointage.xlsx")
    ecrites = gestionnaire.ajouter_lignes(
        [_ligne(), _ligne("MINESEC", 22920)],
        fichier="facture-1.pdf",
        empreinte="abc123",
        utilisateur="poste-1",
        pages=1,
        confiance=0.95,
    )
    assert ecrites == 2

    classeur = load_workbook(gestionnaire.chemin)
    assert FEUILLE_RESUME in classeur.sheetnames
    assert FEUILLE_ANOMALIES in classeur.sheetnames
    assert FEUILLE_JOURNAL in classeur.sheetnames
    assert "MINSANTE" in classeur.sheetnames
    assert "MINESEC" in classeur.sheetnames

    feuille = classeur["MINSANTE"]
    entetes = [cellule.value for cellule in feuille[1]]
    assert entetes == list(COLONNES) + list(COLONNES_TECHNIQUES)
    assert feuille.max_row == 2
    assert feuille.cell(row=2, column=len(COLONNES) + 2).value == "facture-1.pdf"


def test_ajout_successif_et_resume(tmp_path):
    gestionnaire = ExcelManager(tmp_path / "pointage.xlsx")
    gestionnaire.ajouter_lignes([_ligne()], fichier="f1.pdf", empreinte="h1")
    gestionnaire.ajouter_lignes([_ligne(), _ligne()], fichier="f2.pdf", empreinte="h2")

    classeur = load_workbook(gestionnaire.chemin)
    assert classeur["MINSANTE"].max_row == 4  # 1 entête + 3 lignes

    resume = classeur[FEUILLE_RESUME]
    lignes_resume = list(resume.iter_rows(min_row=2, values_only=True))
    assert lignes_resume[0][0] == "MINSANTE"
    assert lignes_resume[0][1] == "mars-2026"
    assert lignes_resume[0][2] == 3                  # nb lignes
    assert lignes_resume[0][4] == 3 * 39200          # montant HT cumulé
    assert lignes_resume[-1][0] == "TOTAL GÉNÉRAL"

    journal = classeur[FEUILLE_JOURNAL]
    assert journal.max_row == 3                      # 1 entête + 2 factures


def test_detection_des_doublons(tmp_path):
    gestionnaire = ExcelManager(tmp_path / "pointage.xlsx")
    gestionnaire.ajouter_lignes([_ligne()], fichier="f1.pdf", empreinte="empreinte-1")

    assert gestionnaire.fichier_deja_traite("empreinte-1") == "f1.pdf"
    assert gestionnaire.fichier_deja_traite("inconnue") is None


def test_anomalies_reportees(tmp_path):
    gestionnaire = ExcelManager(tmp_path / "pointage.xlsx")
    ligne = _ligne(statut=MARQUEUR_A_VERIFIER)
    ligne.anomalies = ["TVA incohérente", "Ministère incertain"]
    gestionnaire.ajouter_lignes([ligne], fichier="f1.pdf", empreinte="h1")

    classeur = load_workbook(gestionnaire.chemin)
    anomalies = classeur[FEUILLE_ANOMALIES]
    assert anomalies.max_row == 3  # 1 entête + 2 anomalies
    assert anomalies.cell(row=2, column=7).value == "TVA incohérente"


def test_ecriture_atomique_conserve_le_fichier(tmp_path, monkeypatch):
    """Si l'enregistrement échoue, le classeur précédent reste intact."""
    gestionnaire = ExcelManager(tmp_path / "pointage.xlsx")
    gestionnaire.ajouter_lignes([_ligne()], fichier="f1.pdf", empreinte="h1")
    taille_avant = gestionnaire.chemin.stat().st_size

    from openpyxl import Workbook

    def sauvegarde_qui_echoue(self, *args, **kwargs):
        raise OSError("disque plein")

    monkeypatch.setattr(Workbook, "save", sauvegarde_qui_echoue)

    try:
        gestionnaire.ajouter_lignes([_ligne()], fichier="f2.pdf", empreinte="h2")
    except Exception:
        pass

    assert gestionnaire.chemin.exists()
    assert gestionnaire.chemin.stat().st_size == taille_avant
    assert not list(tmp_path.glob("*.tmp.xlsx"))


def test_statistiques(tmp_path):
    gestionnaire = ExcelManager(tmp_path / "pointage.xlsx")
    assert gestionnaire.statistiques()["existe"] is False

    gestionnaire.ajouter_lignes([_ligne(), _ligne("MINESEC")], fichier="f1.pdf", empreinte="h1")
    stats = gestionnaire.statistiques()

    assert stats["existe"] is True
    assert stats["lignes"] == 2
    assert stats["factures_traitees"] == 1
    assert {f["nom"] for f in stats["feuilles"]} == {"MINSANTE", "MINESEC"}
